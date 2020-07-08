import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from textblob import TextBlob
import csv
import openpyxl
from openpyxl import load_workbook
import xlrd


path = 'results.xlsx'

#to open the excel sheet
read_data=openpyxl.load_workbook(path)

sheet_obj = read_data.active

tweets = pd.read_csv("results.csv")
tweets.head()

upa={'P':0,'N':0,'O':0}
nda={'P':0,'N':0,'O':0}

def value(ans,x):
    if ans == 1:
        x['P'] += 1
    elif ans == 2:
        x['N'] += 1
    else:
        x['O'] += 1 


max_row=sheet_obj.max_row

infile = 'results.csv'

def senti_ment():
            abc = TextBlob(data)
            if (abc.sentiment.polarity >0 and abc.sentiment.polarity <=1 or abc.sentiment.subjectivity <1 and abc.sentiment.subjectivity >0.7):
                ans=1
                return ans
            elif (abc.sentiment.polarity <0 and abc.sentiment.polarity >= -1 or abc.sentiment.subjectivity >0 and abc.sentiment.subjectivity <0.3):
                ans=2
                return ans
            else:
                ans=3
                return ans

def get_candidate(row):
        parties = []
        text = row[" tidy_tweet"].lower()
        if "upa" in text or "congress" in text or "gandhi" in text:
            parties.append("upa")
        elif "nda" in text or "bjp" or "modi" in text or "amitshah" in text:
            parties.append("nda")
        return ",".join(parties)
      

tweets["party"] = tweets.apply(get_candidate,axis=1)
counts = tweets["party"].value_counts()

#tweets["upa"] = [tweets["party"]=='upa']


for i in range(2,max_row+1):
    data=str(sheet_obj.cell(row=i,column=2).value)
    if 'upa' in data or 'congress' in data or 'gandhi' in data:
                        x=upa
                        ans=senti_ment()
                        value(ans,x)
    elif 'nda' in data or 'bjp' or 'modi' in data or 'amitshah' in data:
                        x=nda
                        ans=senti_ment()
                        value(ans,x)

n=2
positive= (upa['P'],nda['P'])
negative= (upa['N'],nda['N'])
neutral = (upa['O'],nda['O'])

print("::total tweets:::")
print(counts)
print(":::::Sentimenting relevant tweets, After Discarding some tweets (no match with Keywords defined::::: ")
print(upa)
print(nda)

fig, ax = plt.subplots()
index = np.arange(n)

#cursor = mplcursors.cursor(hover=True)
#@cursor.connect("add")
#def on_add(sel):
#    x, y, width, height = sel.artist[sel.target.index].get_bbox().bounds
#    sel.annotation.set(
#        text="{}: {}".format(x + width / 2, height),
#        position=(0, 20))
#    sel.annotation.xy = (x + width / 2, y + height)

plt.text(0.3, 10000, "Keywords for relevant tweets\n For UPA: upa,congress,gandhi \n For NDA: nda,bjp,modi,amitshah", bbox=dict(facecolor='lightblue', alpha=0.5))

ax.set_xticklabels(('UPA','NDA'))
bar_width = 0.3
opacity = 0.8
ax.bar(index, positive, bar_width, alpha=opacity,color='g',label='positive')
ax.bar(index+bar_width, negative, bar_width, alpha=opacity,color='r',label='negative')
ax.bar(index+2*bar_width, neutral, bar_width, alpha=opacity,color='k',label='neutral')

ax.set_xlabel('Alliances',fontsize='16')
ax.set_ylabel('Number of relevant tweets',fontsize='16')
ax.set_xticks(index+bar_width)
plt.title("Tweets analysis between NDA and UPA With Sentiment")
#ax.set_xticklabels(('NDA','UPA'))

#legend(loc='upper left')
plt.legend()

plt.show()
#print (tweets["party"])

#tweets["party"].to_csv(r'NEW.csv', header=True, index=False, sep=' ', mode='a')
